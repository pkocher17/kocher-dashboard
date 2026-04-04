#!/usr/bin/env python3
"""
Build script for The Kocher Team dashboard.
Reads the xlsm/xlsx file and injects data into the template.

Usage:
  python3 build_kocher.py [spreadsheet.xlsm] [--api-key YOUR_KEY] [--pin YOUR_PIN]

Output: index.html (ready to push to GitHub Pages)
"""

import sys, json, re, os
from datetime import datetime

# ── DEFAULTS ──────────────────────────────
SPREADSHEET = "The Kocher Team.xlsm"
TEMPLATE    = "kocher_template.html"
OUTPUT      = "index.html"
GH_OWNER    = "%%GH_OWNER%%"   # filled in by setup
GH_REPO     = "%%GH_REPO%%"    # filled in by setup
MAPS_KEY    = "%%MAPS_API_KEY%%"
PASS_PIN    = "757"

# Parse CLI args
for i, arg in enumerate(sys.argv[1:], 1):
    if arg.endswith(('.xlsx', '.xlsm', '.xls')):
        SPREADSHEET = arg
    elif arg == '--api-key' and i+1 < len(sys.argv):
        MAPS_KEY = sys.argv[i+1]
    elif arg == '--pin' and i+1 < len(sys.argv):
        PASS_PIN = sys.argv[i+1]
    elif arg == '--owner' and i+1 < len(sys.argv):
        GH_OWNER = sys.argv[i+1]
    elif arg == '--repo' and i+1 < len(sys.argv):
        GH_REPO = sys.argv[i+1]

# ── PARSE SPREADSHEET ─────────────────────
def parse_spreadsheet(path):
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl…")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    # Prefer 'Sheet 1 - Agent Single Line', else first sheet
    sheet_name = 'Sheet 1 - Agent Single Line'
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    print(f"Reading sheet: {ws.title}")

    rows_raw = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = -1
    for i, row in enumerate(rows_raw[:10]):
        row_str = [str(c or '') for c in row]
        if any('ML #' in s or s == 'Address' for s in row_str):
            header_row_idx = i
            break
    if header_row_idx == -1:
        print("ERROR: Could not find header row"); sys.exit(1)

    headers = [str(c or '').strip() for c in rows_raw[header_row_idx]]

    def col(name):
        for i, h in enumerate(headers):
            if name in h: return i
        return -1

    iMls   = col('ML #')
    iAddr  = col('Address')
    iSub   = col('Subdivision')
    iSqFt  = col('SqFt')
    iBeds  = col('Beds')
    iBaths = col('Bath')
    iGar   = col('GAR')
    iAcres = col('Acres')
    iPool  = col('Pool')
    iPrice = col('Current Price')
    iClose = col('Close Date') if col('Close Date') >= 0 else col('Closing Date') if col('Closing Date') >= 0 else col('Sold Date')
    iCity  = col('City') if col('City') >= 0 else col('Prop City') if col('Prop City') >= 0 else col('Property City')
    iZip   = col('Zip') if col('Zip') >= 0 else col('ZIP') if col('ZIP') >= 0 else col('Postal')

    parsed = []
    for row in rows_raw[header_row_idx + 1:]:
        if len(row) <= max(iAddr, 0): continue
        addr = str(row[iAddr] or '').replace('  ', ' ').strip() if iAddr >= 0 else ''
        if not addr: continue
        price_raw = str(row[iPrice] or '0').replace('$','').replace(',','').strip() if iPrice >= 0 else '0'
        try:
            price = float(price_raw)
        except ValueError:
            price = 0.0

        # Extract year from close date
        year = 0
        if iClose >= 0 and row[iClose]:
            cd = row[iClose]
            if hasattr(cd, 'year'):   # datetime object from openpyxl
                year = cd.year
            else:
                import re as _re
                m = _re.search(r'\b(20\d{2}|19\d{2})\b', str(cd))
                if m: year = int(m.group(1))

        parsed.append({
            'mls':        str(row[iMls]   or '') if iMls   >= 0 else '',
            'address':    addr,
            'subdivision':str(row[iSub]   or '') if iSub   >= 0 else '',
            'sqft':       str(row[iSqFt]  or '') if iSqFt  >= 0 else '',
            'beds':       str(row[iBeds]  or '') if iBeds  >= 0 else '',
            'baths':      str(row[iBaths] or '') if iBaths >= 0 else '',
            'garage':     str(row[iGar]   or '') if iGar   >= 0 else '',
            'acres':      str(row[iAcres] or '') if iAcres >= 0 else '',
            'pool':       str(row[iPool]  or '') if iPool  >= 0 else '',
            'price':      price,
            'year':       year,
            'city':       str(row[iCity] or '').strip() if iCity >= 0 and row[iCity] else '',
            'zip':        str(row[iZip]  or '').strip() if iZip  >= 0 and row[iZip]  else '',
            'lat': None, 'lng': None
        })

    return parsed


# ── BUILD ─────────────────────────────────
def build():
    if not os.path.exists(SPREADSHEET):
        print(f"Spreadsheet not found: {SPREADSHEET}")
        print("Building with empty data (use Upload Data button in dashboard to add data)")
        data = []
    else:
        data = parse_spreadsheet(SPREADSHEET)
        print(f"Parsed {len(data)} records")

    if not os.path.exists(TEMPLATE):
        print(f"ERROR: Template not found: {TEMPLATE}"); sys.exit(1)

    with open(TEMPLATE, 'r', encoding='utf-8') as f:
        html = f.read()

    now = datetime.now()
    refreshed = now.strftime('%B %-d, %Y at %I:%M %p')

    html = html.replace('%%CLIENT_DATA%%', json.dumps(data, ensure_ascii=False))
    html = html.replace('%%REFRESHED%%', refreshed)
    html = html.replace('%%MAPS_API_KEY%%', MAPS_KEY)
    html = html.replace('%%GH_OWNER%%', GH_OWNER)
    html = html.replace('%%GH_REPO%%', GH_REPO)
    html = html.replace('%%PASS_PIN%%', PASS_PIN)

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"Built {OUTPUT}  ({len(html):,} bytes)")
    return len(data)


if __name__ == '__main__':
    n = build()
    print(f"Done — {n} properties in dashboard")
